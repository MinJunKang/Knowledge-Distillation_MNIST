
# First install xlrd and openpyxl

import Header

try:
    from openpyxl import Workbook
    import pandas as pd

except ImportError as ex:
    print('Please install xlrd and openpyxl to start')
    print(ex)

class Analyzer(object):
    def __init__(self,path_student_txt = "",path_result_excel = ""):
        self.excel_path = path_result_excel
        self.text_path = path_student_txt

        if(Header.os.path.isfile(path_student_txt) == False):
            print("Train the student model first to analyze the result")
            exit(1)
        [self.__data , self.__num_layer] = self.__Load_text_data()
        self.__layer_num = int(self.__Text_to_Excel())

        return
    def __Load_text_data(self):
        f = open(self.text_path,'r')
        lines = f.readlines()
        data = []
        num_layer = []
        i = 1
        for line in lines:
            num = line.count(' ')
            item = line.split(" ")
            if (i-3)%6 == 0:
                num_layer.append(num)
                num_layer.append(num)
            if (i-5)%6 == 0:
                accuracy = item[14]
                data.append(accuracy)
            if  i % 6 == 0:
                parameter = item[4]
                data.append(parameter)
            i = i+1
        return data, num_layer

    def __Text_to_Excel(self):

        # Setting
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "first_sheet"
        title = ["accuracy","num_parameter"]

        # Calculate layer number and layer_row_pos
        layer_num = int(Header.np.max(self.__num_layer) / 2) -1
        layer_row_pos = Header.np.ones(layer_num)

        k = 0
        for col in range(1,layer_num*2+1):
            ws1.cell(row = 1, column = col, value = title[int(k%2)]+'%d'%(int(k/2)))
            k = k +1
        
        i = 0
        while(1):
            layer = int(self.__num_layer[i] / 2) -1
            for col in range(2 * layer - 1 ,2 * layer + 1):
                ws1.cell(row = layer_row_pos[layer-1]+1, column = col, value = self.__data[i])
                i = i+1
            layer_row_pos[layer-1] += 1
            if i == len(self.__num_layer):
                break

        # Save the excel file
        wb.save(self.excel_path)

        return layer_num

    def Excelgraph(self,num_teacher_params):
        data = pd.read_excel(self.excel_path)
        data.head()
        temp = []
        temp2 = []
        for i in range(self.__layer_num):
            temp.append('num_parameter'+str(i))
            temp2.append('accuracy'+str(i))
        Header.plt.figure()
        for i in range(self.__layer_num):
            Header.plt.scatter(100*(data[temp[i]]/num_teacher_params), data[temp2[i]], label = 'layer%i' %(int(i)+1),s=5)
        Header.plt.xlabel("Complexity(# Student params / # Teacher params)[%]")
        Header.plt.ylabel("Accuracy")
        Header.plt.title("Complexity - Accuracy Result")
        Header.plt.legend()
        Header.plt.show()
